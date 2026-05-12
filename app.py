"""
VT x MADUP US/UK 광고 성과 대시보드
- 데이터 소스: Google Sheets ("2026 RD" 시트)
- 메인 지표: CTR, ROAS, 지출 금액
"""
from __future__ import annotations

import calendar
import datetime
import io
import time
from datetime import date, timedelta
from pathlib import Path

import pandas as pd
import requests
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
import streamlit.components.v1 as components

SHEET_ID = "1mFrK6lm-NZCd6_uHGeEzPYwywXXc1XfBIURbbmLXuLc"
SHEET_GID = "1759920788"
CSV_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=csv&gid={SHEET_GID}"
LOCAL_CSV = Path(__file__).parent / "rd_data.csv"

# ---------- Page ----------
st.set_page_config(
    page_title="VT × MADUP 광고 성과 대시보드",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown(
    """
    <style>
    .block-container { padding-top: 1.5rem; padding-bottom: 2rem; max-width: 1400px;}
    [data-testid="stMetric"] {
        background: #ffffff;
        border: 1px solid #e6e8eb;
        border-radius: 12px;
        padding: 16px 18px;
        box-shadow: 0 1px 2px rgba(15,23,42,0.04);
    }
    [data-testid="stMetricLabel"] p { font-size: 0.85rem; color:#64748b; font-weight:500;}
    [data-testid="stMetricValue"] { font-size: 1.8rem; font-weight: 700; color:#0f172a;}
    h1 { font-size: 1.7rem !important; font-weight: 700 !important; color:#0f172a;}
    h2 { font-size: 1.2rem !important; font-weight: 600 !important; color:#1e293b; margin-top: 1.2rem !important;}
    h3 { font-size: 1.05rem !important; font-weight: 600 !important; color:#334155;}
    .stTabs [data-baseweb="tab-list"] { gap: 4px; }
    .stTabs [data-baseweb="tab"] {
        background: #f1f5f9;
        border-radius: 8px 8px 0 0;
        padding: 8px 18px;
        font-weight: 500;
    }
    .stTabs [aria-selected="true"] { background: #0ea5e9; color: white; }
    .caption-muted { color: #64748b; font-size: 0.85rem; }
    </style>
    """,
    unsafe_allow_html=True,
)

# ---------- Cleanup any prior floating clock (after a hot reload) ----------
components.html(
    """
<script>
try {
  const P = window.parent.document;
  const old = P.getElementById('madup-clocks');
  if (old) old.remove();
  if (window.parent.__madupClockTimer) {
    clearInterval(window.parent.__madupClockTimer);
    window.parent.__madupClockTimer = null;
  }
} catch (e) {}
</script>
    """,
    height=0,
)


# ---------- Data ----------
def _parse_num(x) -> float:
    if pd.isna(x):
        return 0.0
    s = str(x).replace("₩", "").replace(",", "").replace(" ", "").strip()
    if s in {"", "-", "#REF!", "#VALUE!", "#DIV/0!"}:
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def _fetch_csv_bytes(url: str, attempts: int = 3) -> bytes:
    """Stream the Sheets export with generous timeouts and retries."""
    last_err = None
    for i in range(attempts):
        try:
            with requests.get(url, timeout=(15, 300), stream=True) as r:
                r.raise_for_status()
                buf = io.BytesIO()
                for chunk in r.iter_content(chunk_size=256 * 1024):
                    if chunk:
                        buf.write(chunk)
                return buf.getvalue()
        except (requests.RequestException, OSError) as e:
            last_err = e
            time.sleep(2 ** i)
    raise RuntimeError(f"Google Sheets CSV 다운로드 실패: {last_err}")


@st.cache_data(show_spinner=False)
def load_data(refresh_key: int = 0) -> pd.DataFrame:
    """Load and clean the RD sheet. refresh_key forces cache invalidation."""
    if LOCAL_CSV.exists() and refresh_key == 0:
        raw = LOCAL_CSV.read_bytes()
    else:
        raw = _fetch_csv_bytes(CSV_URL)
        try:
            LOCAL_CSV.write_bytes(raw)
        except OSError:
            pass  # read-only FS는 무시

    df = pd.read_csv(io.BytesIO(raw), header=2, low_memory=False)

    # Keep only rows with valid date and country
    df["일별"] = pd.to_datetime(df["일별"], errors="coerce")
    df = df.dropna(subset=["일별"]).copy()
    df = df[df["국가"].notna()]

    # Numeric conversions (cost = NET, "비용" 컬럼)
    df["cost"] = df["비용"].apply(_parse_num)
    df["cost_gross"] = df["지출 금액(GROSS)"].apply(_parse_num)
    df["impressions"] = df["노출수"].apply(_parse_num)
    df["clicks"] = df["클릭수(목적지)"].apply(_parse_num)
    df["aa_sales_usd"] = pd.to_numeric(df["AA Total Sales"], errors="coerce").fillna(0)
    df["ga4_revenue_usd"] = pd.to_numeric(df["GA4 총 구매 수익(USD)"], errors="coerce").fillna(0)
    df["ga4_purchases"] = pd.to_numeric(df["GA4 구매"], errors="coerce").fillna(0)
    df["aa_purchases"] = pd.to_numeric(df["AA purchase"], errors="coerce").fillna(0)

    df["country"] = df["국가"].astype(str).str.strip()
    df["media"] = df["매체"].astype(str).str.strip()
    df["item"] = df["품목"].astype(str).str.strip()
    df["objective"] = df["캠페인구분"].astype(str).str.strip()
    df["campaign"] = df["캠페인 이름"].astype(str).str.strip()

    return df


def kpi(df: pd.DataFrame, fx_rate: float) -> dict:
    cost = df["cost"].sum()
    imp = df["impressions"].sum()
    clk = df["clicks"].sum()
    sales_usd = df["aa_sales_usd"].sum()
    sales_krw = sales_usd * fx_rate
    purchases = df["aa_purchases"].sum() + df["ga4_purchases"].sum()
    return {
        "cost_krw": cost,
        "impressions": imp,
        "clicks": clk,
        "ctr": (clk / imp * 100) if imp > 0 else 0,
        "sales_usd": sales_usd,
        "sales_krw": sales_krw,
        "roas": (sales_krw / cost) if cost > 0 else 0,
        "cpc": (cost / clk) if clk > 0 else 0,
        "cpm": (cost / imp * 1000) if imp > 0 else 0,
        "purchases": purchases,
    }


def agg_by(df: pd.DataFrame, group_cols, fx_rate: float) -> pd.DataFrame:
    g = df.groupby(group_cols, dropna=False).agg(
        cost=("cost", "sum"),
        impressions=("impressions", "sum"),
        clicks=("clicks", "sum"),
        sales_usd=("aa_sales_usd", "sum"),
        purchases=("aa_purchases", "sum"),
    ).reset_index()
    g["sales_krw"] = g["sales_usd"] * fx_rate
    g["ctr"] = (g["clicks"] / g["impressions"]).fillna(0) * 100
    g["roas"] = (g["sales_krw"] / g["cost"]).fillna(0)
    g["cpc"] = (g["cost"] / g["clicks"]).replace([float("inf")], 0).fillna(0)
    g["cpm"] = (g["cost"] / g["impressions"] * 1000).replace([float("inf")], 0).fillna(0)
    g.loc[g["impressions"] == 0, "ctr"] = 0
    g.loc[g["cost"] == 0, "roas"] = 0
    return g


def fmt_krw(v: float) -> str:
    return f"₩{v:,.0f}"


def fmt_pct(v: float) -> str:
    return f"{v:.2f}%"


def fmt_num(v: float) -> str:
    return f"{v:,.0f}"


# ---------- Sidebar ----------
with st.sidebar:
    st.markdown("### ⚙️ 설정")

    if "refresh_key" not in st.session_state:
        st.session_state.refresh_key = 0

    if st.button("🔄 시트에서 최신 데이터 새로고침", width="stretch"):
        st.session_state.refresh_key += 1
        load_data.clear()
        st.rerun()

    fx_rate = st.number_input(
        "환율 (1 USD → KRW)",
        min_value=500.0,
        max_value=3000.0,
        value=1470.0,
        step=10.0,
        help="AA Total Sales는 USD 단위입니다. ROAS는 (USD 매출 × 환율) / NET 지출로 계산.",
    )

with st.spinner("데이터를 불러오는 중..."):
    df_all = load_data(refresh_key=st.session_state.get("refresh_key", 0))

with st.sidebar:
    st.markdown("---")
    st.markdown("### 🔍 필터")

    date_min = df_all["일별"].min().date()
    date_max = df_all["일별"].max().date()

    # 기본값: 이번 달 (데이터 가용 범위 내에서)
    _today = datetime.date.today()
    _m_start = _today.replace(day=1)
    _last_day = calendar.monthrange(_today.year, _today.month)[1]
    _m_end = _today.replace(day=_last_day)
    _def_start = max(_m_start, date_min)
    _def_end = min(_m_end, date_max)
    if _def_start > _def_end:  # 이번 달에 데이터 없으면 가용 마지막 달로 폴백
        _def_end = date_max
        _def_start = max(date_max.replace(day=1), date_min)

    date_range = st.date_input(
        "기간 (기본: 이번 달)",
        value=(_def_start, _def_end),
        min_value=date_min,
        max_value=date_max,
    )
    if isinstance(date_range, (list, tuple)) and len(date_range) == 2:
        d_start, d_end = date_range
    else:
        d_start, d_end = date_min, date_max

    # 비교 기간 (기본: 직전 동일 길이 구간, 사용자 수정 가능)
    _auto_days = (d_end - d_start).days + 1
    _auto_prev_end = d_start - timedelta(days=1)
    _auto_prev_start = _auto_prev_end - timedelta(days=_auto_days - 1)
    _auto_prev_start = max(_auto_prev_start, date_min)
    _auto_prev_end = max(_auto_prev_end, _auto_prev_start)
    cmp_range = st.date_input(
        "비교 기간 (기본: 직전 동일 길이)",
        value=(_auto_prev_start, _auto_prev_end),
        min_value=date_min,
        max_value=date_max,
        help="KPI 카드 아래 '+/-% vs 직전' 계산에 사용됩니다.",
    )
    if isinstance(cmp_range, (list, tuple)) and len(cmp_range) == 2:
        cmp_start, cmp_end = cmp_range
    else:
        cmp_start, cmp_end = _auto_prev_start, _auto_prev_end

    def _unique_str(col: str) -> list[str]:
        s = df_all[col].dropna().astype(str).str.strip()
        s = s[(s != "") & (s.str.lower() != "nan")]
        return sorted(s.unique().tolist())

    countries = _unique_str("country")
    medias = _unique_str("media")
    items = _unique_str("item")
    objectives = _unique_str("objective")

    sel_countries = st.multiselect("국가", countries, default=countries)
    sel_medias = st.multiselect("매체", medias, default=medias)
    sel_items = st.multiselect("품목", items, default=items)
    sel_obj = st.multiselect("캠페인 목적", objectives, default=objectives)

# Apply filters
mask = (
    (df_all["일별"].dt.date >= d_start)
    & (df_all["일별"].dt.date <= d_end)
    & (df_all["country"].isin(sel_countries))
    & (df_all["media"].isin(sel_medias))
    & (df_all["item"].isin(sel_items))
    & (df_all["objective"].isin(sel_obj))
)
df = df_all[mask].copy()


# 비교 기간 (사이드바에서 사용자 지정)
_prev_start, _prev_end = cmp_start, cmp_end
_prev_mask = (
    (df_all["일별"].dt.date >= _prev_start)
    & (df_all["일별"].dt.date <= _prev_end)
    & (df_all["country"].isin(sel_countries))
    & (df_all["media"].isin(sel_medias))
    & (df_all["item"].isin(sel_items))
    & (df_all["objective"].isin(sel_obj))
)
df_prev = df_all[_prev_mask].copy()


def _pct_delta(cur, prev) -> str | None:
    if prev is None or pd.isna(prev) or prev == 0:
        return None
    return f"{(cur - prev) / prev * 100:+.1f}% vs 직전"


# ---------- Header ----------
components.html(
    """
<div id="madup-clocks-inline" style="
  display:inline-flex; gap:14px; align-items:center;
  background:rgba(15,23,42,0.92); color:#f8fafc;
  padding:8px 14px; border-radius:10px;
  font-family:ui-monospace,SFMono-Regular,Menlo,monospace;
  font-size:12.5px; font-weight:500; letter-spacing:0.2px;
  box-shadow:0 4px 12px rgba(15,23,42,0.18);
"></div>
<script>
(function () {
  function block(flag, label, tz) {
    const opts = {timeZone: tz, hour:'2-digit', minute:'2-digit', second:'2-digit', hour12:false};
    const t = new Date().toLocaleTimeString('en-GB', opts);
    const d = new Date().toLocaleDateString('en-CA', {timeZone: tz});
    return '<span style="display:flex;flex-direction:column;line-height:1.15;">'
      + '<span style="opacity:.7;font-size:10.5px;">' + flag + ' ' + label + '</span>'
      + '<span>' + t + '</span>'
      + '<span style="opacity:.45;font-size:9.5px;">' + d + '</span>'
      + '</span>';
  }
  function divider() {
    return '<span style="width:1px;height:30px;background:rgba(255,255,255,0.18);"></span>';
  }
  function update() {
    const el = document.getElementById('madup-clocks-inline');
    if (!el) return;
    el.innerHTML = block('🇰🇷','KST','Asia/Seoul')
      + divider()
      + block('🇺🇸','PT (LA)','America/Los_Angeles');
  }
  update();
  setInterval(update, 1000);
})();
</script>
    """,
    height=64,
)

st.markdown("# 📊 VT × MADUP 광고 성과 대시보드")
st.markdown(
    f"<div class='caption-muted'>기간: <b>{d_start}</b> ~ <b>{d_end}</b>  ·  "
    f"비교 기간: <b>{_prev_start}</b> ~ <b>{_prev_end}</b>  ·  "
    f"국가 {len(sel_countries)}개 · 매체 {len(sel_medias)}개 · 품목 {len(sel_items)}개  ·  "
    f"적용 환율: 1 USD = ₩{fx_rate:,.0f}</div>",
    unsafe_allow_html=True,
)
st.markdown("")

if df.empty:
    st.warning("선택한 조건에 해당하는 데이터가 없습니다. 필터를 조정해 주세요.")
    st.stop()


# ---------- KPI ----------
k = kpi(df, fx_rate)
k_prev = kpi(df_prev, fx_rate)

c1, c2, c3, c4, c5 = st.columns(5)
c1.metric("💸 지출 금액 (NET)", fmt_krw(k["cost_krw"]),
          _pct_delta(k["cost_krw"], k_prev["cost_krw"]))
c2.metric("🛒 매출 (USD)", f"${k['sales_usd']:,.0f}",
          _pct_delta(k["sales_usd"], k_prev["sales_usd"]),
          help=f"≈ {fmt_krw(k['sales_krw'])}")
c3.metric("📈 ROAS", f"{k['roas']:.2f}",
          _pct_delta(k["roas"], k_prev["roas"]),
          help="매출(KRW 환산) / NET 지출(KRW)")
c4.metric("🎯 CTR", fmt_pct(k["ctr"]),
          _pct_delta(k["ctr"], k_prev["ctr"]))
c5.metric("👀 노출수", f"{k['impressions']/1e6:.2f}M",
          _pct_delta(k["impressions"], k_prev["impressions"]))

c6, c7, c8, c9 = st.columns(4)
c6.metric("🖱️ 클릭수", fmt_num(k["clicks"]),
          _pct_delta(k["clicks"], k_prev["clicks"]))
c7.metric("CPC", fmt_krw(k["cpc"]),
          _pct_delta(k["cpc"], k_prev["cpc"]),
          delta_color="inverse")
c8.metric("CPM", fmt_krw(k["cpm"]),
          _pct_delta(k["cpm"], k_prev["cpm"]),
          delta_color="inverse")
c9.metric("🛍️ 총 구매 건수", fmt_num(k["purchases"]),
          _pct_delta(k["purchases"], k_prev["purchases"]))


# ---------- Tabs ----------
tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs(
    ["🌏 국가별", "📡 매체별", "📦 품목별", "📅 추세", "🧾 상세 데이터", "🏆 우수 소재"]
)

# === Tab 1: Country ===
with tab1:
    country = agg_by(df, ["country"], fx_rate).sort_values("cost", ascending=False)

    st.markdown("### 국가별 핵심 성과")
    cc1, cc2 = st.columns([2, 3])

    with cc1:
        fig_pie = px.pie(
            country, names="country", values="cost", hole=0.55,
            title="지출 비중 (국가별)",
            color_discrete_sequence=px.colors.qualitative.Set2,
        )
        fig_pie.update_traces(textinfo="percent+label")
        fig_pie.update_layout(height=380, margin=dict(t=50, b=10, l=10, r=10), showlegend=False)
        st.plotly_chart(fig_pie, width="stretch")

    with cc2:
        fig_bar = go.Figure()
        fig_bar.add_trace(go.Bar(
            x=country["country"], y=country["ctr"], name="CTR (%)",
            marker_color="#0ea5e9", yaxis="y",
            text=[f"{v:.2f}%" for v in country["ctr"]], textposition="outside",
        ))
        fig_bar.add_trace(go.Scatter(
            x=country["country"], y=country["roas"], name="ROAS",
            mode="lines+markers+text", marker=dict(size=10, color="#f97316"),
            line=dict(color="#f97316", width=3), yaxis="y2",
            text=[f"{v:.2f}" for v in country["roas"]], textposition="top center",
        ))
        fig_bar.update_layout(
            title="국가별 CTR & ROAS",
            height=380,
            margin=dict(t=50, b=10, l=10, r=10),
            yaxis=dict(title="CTR (%)", side="left"),
            yaxis2=dict(title="ROAS", side="right", overlaying="y"),
            legend=dict(orientation="h", y=-0.15, x=0.5, xanchor="center"),
        )
        st.plotly_chart(fig_bar, width="stretch")

    st.markdown("### 국가별 상세")
    show = country.rename(columns={
        "country": "국가", "cost": "지출(NET)", "impressions": "노출수",
        "clicks": "클릭수", "ctr": "CTR(%)", "sales_usd": "매출(USD)",
        "sales_krw": "매출(KRW환산)", "roas": "ROAS", "cpc": "CPC", "cpm": "CPM",
        "purchases": "구매",
    })
    st.dataframe(
        show.style.format({
            "지출(NET)": "₩{:,.0f}", "노출수": "{:,.0f}", "클릭수": "{:,.0f}",
            "CTR(%)": "{:.2f}%", "매출(USD)": "${:,.2f}",
            "매출(KRW환산)": "₩{:,.0f}", "ROAS": "{:.2f}",
            "CPC": "₩{:,.0f}", "CPM": "₩{:,.0f}", "구매": "{:,.0f}",
        }),
        width="stretch", hide_index=True,
    )

    st.markdown("### 국가 × 매체 매트릭스")
    pivot = agg_by(df, ["country", "media"], fx_rate)
    if not pivot.empty:
        cm1, cm2 = st.columns(2)
        with cm1:
            heat_ctr = pivot.pivot_table(index="country", columns="media", values="ctr", aggfunc="sum").fillna(0)
            fig_h1 = px.imshow(heat_ctr, text_auto=".2f", aspect="auto",
                              color_continuous_scale="Blues", title="CTR(%) — 국가 × 매체")
            fig_h1.update_layout(height=320, margin=dict(t=50, b=10, l=10, r=10))
            st.plotly_chart(fig_h1, width="stretch")
        with cm2:
            heat_roas = pivot.pivot_table(index="country", columns="media", values="roas", aggfunc="sum").fillna(0)
            fig_h2 = px.imshow(heat_roas, text_auto=".2f", aspect="auto",
                              color_continuous_scale="Oranges", title="ROAS — 국가 × 매체")
            fig_h2.update_layout(height=320, margin=dict(t=50, b=10, l=10, r=10))
            st.plotly_chart(fig_h2, width="stretch")


# === Tab 2: Media ===
with tab2:
    media = agg_by(df, ["media"], fx_rate).sort_values("cost", ascending=False)

    st.markdown("### 매체별 핵심 성과")
    mm1, mm2, mm3 = st.columns(3)
    mm1.metric("최다 지출", media.iloc[0]["media"], fmt_krw(media.iloc[0]["cost"]))
    best_ctr = media.sort_values("ctr", ascending=False).iloc[0]
    mm2.metric("최고 CTR", best_ctr["media"], fmt_pct(best_ctr["ctr"]))
    best_roas = media.sort_values("roas", ascending=False).iloc[0]
    mm3.metric("최고 ROAS", best_roas["media"], f"{best_roas['roas']:.2f}")

    m1, m2 = st.columns(2)
    with m1:
        fig = px.bar(
            media, x="media", y="cost",
            title="매체별 지출", text_auto=".2s",
            color="media", color_discrete_sequence=px.colors.qualitative.Pastel,
        )
        fig.update_layout(height=350, showlegend=False, margin=dict(t=50, b=10, l=10, r=10))
        st.plotly_chart(fig, width="stretch")
    with m2:
        fig = go.Figure()
        fig.add_trace(go.Bar(x=media["media"], y=media["ctr"], name="CTR(%)",
                              marker_color="#0ea5e9",
                              text=[f"{v:.2f}%" for v in media["ctr"]], textposition="outside"))
        fig.add_trace(go.Scatter(x=media["media"], y=media["roas"], name="ROAS",
                                  mode="lines+markers", marker=dict(size=12, color="#f97316"),
                                  line=dict(color="#f97316", width=3), yaxis="y2"))
        fig.update_layout(
            title="매체별 CTR & ROAS", height=350,
            yaxis=dict(title="CTR(%)"),
            yaxis2=dict(title="ROAS", side="right", overlaying="y"),
            margin=dict(t=50, b=10, l=10, r=10),
            legend=dict(orientation="h", y=-0.2, x=0.5, xanchor="center"),
        )
        st.plotly_chart(fig, width="stretch")

    st.markdown("### 매체 × 국가 지출 분포")
    media_country = agg_by(df, ["media", "country"], fx_rate)
    fig = px.bar(
        media_country, x="media", y="cost", color="country",
        text_auto=".2s", barmode="stack",
        color_discrete_sequence=px.colors.qualitative.Set2,
    )
    fig.update_layout(height=380, margin=dict(t=30, b=10, l=10, r=10))
    st.plotly_chart(fig, width="stretch")

    st.markdown("### 매체별 상세")
    show = media.rename(columns={
        "media": "매체", "cost": "지출(NET)", "impressions": "노출수",
        "clicks": "클릭수", "ctr": "CTR(%)", "sales_usd": "매출(USD)",
        "sales_krw": "매출(KRW환산)", "roas": "ROAS", "cpc": "CPC", "cpm": "CPM",
        "purchases": "구매",
    })
    st.dataframe(
        show.style.format({
            "지출(NET)": "₩{:,.0f}", "노출수": "{:,.0f}", "클릭수": "{:,.0f}",
            "CTR(%)": "{:.2f}%", "매출(USD)": "${:,.2f}",
            "매출(KRW환산)": "₩{:,.0f}", "ROAS": "{:.2f}",
            "CPC": "₩{:,.0f}", "CPM": "₩{:,.0f}", "구매": "{:,.0f}",
        }),
        width="stretch", hide_index=True,
    )


# === Tab 3: Item ===
with tab3:
    item = agg_by(df, ["item"], fx_rate).sort_values("cost", ascending=False)

    st.markdown("### 품목별 핵심 성과")
    i1, i2 = st.columns(2)
    with i1:
        fig = px.bar(
            item, x="item", y="cost", text_auto=".2s",
            title="품목별 지출",
            color="cost", color_continuous_scale="Blues",
        )
        fig.update_layout(height=380, margin=dict(t=50, b=10, l=10, r=10), coloraxis_showscale=False)
        st.plotly_chart(fig, width="stretch")
    with i2:
        fig = go.Figure()
        fig.add_trace(go.Bar(x=item["item"], y=item["ctr"], name="CTR(%)",
                              marker_color="#0ea5e9",
                              text=[f"{v:.2f}%" for v in item["ctr"]], textposition="outside"))
        fig.add_trace(go.Scatter(x=item["item"], y=item["roas"], name="ROAS",
                                  mode="lines+markers", marker=dict(size=10, color="#f97316"),
                                  line=dict(color="#f97316", width=3), yaxis="y2"))
        fig.update_layout(
            title="품목별 CTR & ROAS", height=380,
            yaxis=dict(title="CTR(%)"),
            yaxis2=dict(title="ROAS", side="right", overlaying="y"),
            margin=dict(t=50, b=10, l=10, r=10),
            legend=dict(orientation="h", y=-0.2, x=0.5, xanchor="center"),
        )
        st.plotly_chart(fig, width="stretch")

    st.markdown("### 품목 × 국가 성과")
    ic = agg_by(df, ["item", "country"], fx_rate)
    ic1, ic2 = st.columns(2)
    with ic1:
        heat = ic.pivot_table(index="item", columns="country", values="cost", aggfunc="sum").fillna(0)
        fig = px.imshow(heat, text_auto=".2s", aspect="auto",
                       color_continuous_scale="Blues", title="지출(KRW) — 품목 × 국가")
        fig.update_layout(height=380, margin=dict(t=50, b=10, l=10, r=10))
        st.plotly_chart(fig, width="stretch")
    with ic2:
        heat = ic.pivot_table(index="item", columns="country", values="roas", aggfunc="sum").fillna(0)
        fig = px.imshow(heat, text_auto=".2f", aspect="auto",
                       color_continuous_scale="Oranges", title="ROAS — 품목 × 국가")
        fig.update_layout(height=380, margin=dict(t=50, b=10, l=10, r=10))
        st.plotly_chart(fig, width="stretch")

    st.markdown("### 품목별 상세")
    show = item.rename(columns={
        "item": "품목", "cost": "지출(NET)", "impressions": "노출수",
        "clicks": "클릭수", "ctr": "CTR(%)", "sales_usd": "매출(USD)",
        "sales_krw": "매출(KRW환산)", "roas": "ROAS", "cpc": "CPC", "cpm": "CPM",
        "purchases": "구매",
    })
    st.dataframe(
        show.style.format({
            "지출(NET)": "₩{:,.0f}", "노출수": "{:,.0f}", "클릭수": "{:,.0f}",
            "CTR(%)": "{:.2f}%", "매출(USD)": "${:,.2f}",
            "매출(KRW환산)": "₩{:,.0f}", "ROAS": "{:.2f}",
            "CPC": "₩{:,.0f}", "CPM": "₩{:,.0f}", "구매": "{:,.0f}",
        }),
        width="stretch", hide_index=True,
    )


# === Tab 4: Trend ===
with tab4:
    daily = agg_by(df, ["일별"], fx_rate).sort_values("일별")
    daily["일별"] = pd.to_datetime(daily["일별"])

    st.markdown("### 일자별 추세")
    t1, t2 = st.columns(2)
    with t1:
        fig = px.area(daily, x="일별", y="cost", title="일자별 지출(KRW)",
                     color_discrete_sequence=["#0ea5e9"])
        fig.update_layout(height=320, margin=dict(t=50, b=10, l=10, r=10))
        st.plotly_chart(fig, width="stretch")
    with t2:
        fig = go.Figure()
        fig.add_trace(go.Scatter(x=daily["일별"], y=daily["ctr"], name="CTR(%)",
                                  mode="lines", line=dict(color="#0ea5e9", width=2.5)))
        fig.add_trace(go.Scatter(x=daily["일별"], y=daily["roas"], name="ROAS",
                                  mode="lines", line=dict(color="#f97316", width=2.5), yaxis="y2"))
        fig.update_layout(
            title="일자별 CTR & ROAS", height=320,
            yaxis=dict(title="CTR(%)"),
            yaxis2=dict(title="ROAS", side="right", overlaying="y"),
            margin=dict(t=50, b=10, l=10, r=10),
            legend=dict(orientation="h", y=-0.2, x=0.5, xanchor="center"),
        )
        st.plotly_chart(fig, width="stretch")

    st.markdown("### 국가별 일자 추세 (지출)")
    daily_country = agg_by(df, ["일별", "country"], fx_rate).sort_values("일별")
    daily_country["일별"] = pd.to_datetime(daily_country["일별"])
    fig = px.line(daily_country, x="일별", y="cost", color="country",
                 color_discrete_sequence=px.colors.qualitative.Set2)
    fig.update_layout(height=380, margin=dict(t=30, b=10, l=10, r=10),
                      yaxis_title="지출(NET)", xaxis_title="")
    st.plotly_chart(fig, width="stretch")

    st.markdown("### 매체별 일자 추세 (CTR)")
    daily_media = agg_by(df, ["일별", "media"], fx_rate).sort_values("일별")
    daily_media["일별"] = pd.to_datetime(daily_media["일별"])
    fig = px.line(daily_media, x="일별", y="ctr", color="media",
                 color_discrete_sequence=px.colors.qualitative.Pastel)
    fig.update_layout(height=380, margin=dict(t=30, b=10, l=10, r=10),
                      yaxis_title="CTR(%)", xaxis_title="")
    st.plotly_chart(fig, width="stretch")


# === Tab 5: Raw ===
with tab5:
    st.markdown("### 캠페인 단위 상세 데이터")
    camp = agg_by(df, ["country", "media", "item", "campaign"], fx_rate).sort_values("cost", ascending=False)
    show = camp.rename(columns={
        "country": "국가", "media": "매체", "item": "품목", "campaign": "캠페인",
        "cost": "지출(NET)", "impressions": "노출수", "clicks": "클릭수",
        "ctr": "CTR(%)", "sales_usd": "매출(USD)", "sales_krw": "매출(KRW환산)",
        "roas": "ROAS", "cpc": "CPC", "cpm": "CPM", "purchases": "구매",
    })
    st.dataframe(
        show.style.format({
            "지출(NET)": "₩{:,.0f}", "노출수": "{:,.0f}", "클릭수": "{:,.0f}",
            "CTR(%)": "{:.2f}%", "매출(USD)": "${:,.2f}",
            "매출(KRW환산)": "₩{:,.0f}", "ROAS": "{:.2f}",
            "CPC": "₩{:,.0f}", "CPM": "₩{:,.0f}", "구매": "{:,.0f}",
        }),
        width="stretch", hide_index=True, height=600,
    )

    csv_bytes = show.to_csv(index=False).encode("utf-8-sig")
    st.download_button(
        "⬇️ CSV 다운로드 (필터 적용된 캠페인 단위)",
        csv_bytes, file_name="campaign_performance.csv", mime="text/csv",
    )


# ---------- 우수 소재 (Top 10) helper functions ----------
def _top_creatives(
    df_src: pd.DataFrame,
    d_start, d_end,
    country: str, media: str, objective: str,
    min_cost: float,
    sort_col: str,
    fx_rate: float,
    n: int = 10,
    item_filter: list | None = None,
) -> pd.DataFrame:
    m = (
        (df_src["일별"].dt.date >= d_start)
        & (df_src["일별"].dt.date <= d_end)
        & (df_src["country"] == country)
        & (df_src["media"] == media)
        & (df_src["objective"] == objective)
    )
    sub = df_src[m]
    if item_filter is not None and len(item_filter) > 0:
        sub = sub[sub["item"].isin(item_filter)]
    if sub.empty:
        return sub
    _first_str = lambda s: s.dropna().astype(str).iloc[0] if s.dropna().size else ""
    g = sub.groupby("광고 이름", dropna=False).agg(
        item=("item", _first_str),
        media_=("media", _first_str),
        ad_code=("광고 코드", _first_str),
        aa_link=("AA Attribution tags", _first_str),
        link=("소재 링크", _first_str),
        cost=("cost", "sum"),
        impressions=("impressions", "sum"),
        clicks=("clicks", "sum"),
        sales_usd=("aa_sales_usd", "sum"),
        purchases=("aa_purchases", "sum"),
    ).reset_index()
    g = g[g["cost"] >= min_cost].copy()
    g["ctr"] = (g["clicks"] / g["impressions"]).replace([float("inf")], 0).fillna(0) * 100
    g.loc[g["impressions"] == 0, "ctr"] = 0
    g["roas"] = (g["sales_usd"] * fx_rate / g["cost"]).replace([float("inf")], 0).fillna(0)
    g.loc[g["cost"] == 0, "roas"] = 0
    g = g.sort_values(sort_col, ascending=False).head(n)
    g.insert(0, "순위", range(1, len(g) + 1))
    return g


_TOP_RENAME = {
    "item": "품목",
    "link": "소재 링크",
    "cost": "지출(NET)",
    "impressions": "노출수",
    "clicks": "클릭수",
    "ctr": "CTR(%)",
    "sales_usd": "AA Sales(USD)",
    "roas": "ROAS",
}
_TOP_COLS = ["순위", "품목", "광고 이름", "소재 링크", "지출(NET)", "노출수", "클릭수", "CTR(%)", "AA Sales(USD)", "ROAS"]


def _format_top(df_top: pd.DataFrame) -> pd.DataFrame:
    return df_top.rename(columns=_TOP_RENAME)[_TOP_COLS]


def _render_top(df_top: pd.DataFrame):
    if df_top.empty:
        st.info("조건에 해당하는 소재가 없습니다.")
        return
    show = _format_top(df_top)
    st.dataframe(
        show.style.format({
            "지출(NET)": "₩{:,.0f}",
            "노출수": "{:,.0f}",
            "클릭수": "{:,.0f}",
            "CTR(%)": "{:.2f}%",
            "AA Sales(USD)": "${:,.2f}",
            "ROAS": "{:.2f}",
        }),
        width="stretch",
        hide_index=True,
        column_config={
            "소재 링크": st.column_config.LinkColumn("소재 링크", display_text="🔗 열기"),
            "광고 이름": st.column_config.TextColumn("광고 이름", width="large"),
        },
        height=420,
    )


_XLSX_COLS = [
    ("순위", 6, None),
    ("광고 이름", 60, "@"),
    ("제품", 24, "@"),
    ("매체", 10, "@"),
    ("광고 코드", 50, "@"),
    ("AA 링크", 60, "@"),
    ("지출 금액 (₩)", 16, '₩#,##0'),
    ("노출", 12, '#,##0'),
    ("클릭", 10, '#,##0'),
    ("CTR", 9, '0.00%'),
    ("AA Total Sales ($)", 22, '$#,##0.00'),
    ("AA 구매", 10, '0'),
    ("ROAS", 10, '0.00%'),
]


def _build_top10_xlsx(df_all_, d_start_, d_end_, fx_rate_, items_to_include: list) -> bytes:
    """우수 소재 Top 10 (CTR + Sales) 을 품목별로 묶어 보고서 형태 xlsx로 반환."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    HEADER_FILL = PatternFill("solid", fgColor="305496")
    HEADER_FONT = Font(color="FFFFFF", bold=True)
    CENTER = Alignment(horizontal="center", vertical="center")
    LINK_FONT = Font(color="0563C1", underline="single")
    TITLE_FONT = Font(bold=True, size=12)
    SUB_FONT = Font(bold=True, size=10, color="595959")
    GROUP_FONT = Font(bold=True, size=12)
    SECTION_FONT = Font(bold=True, size=10, color="595959")
    BORDER = Border(*(Side(style="thin", color="D9D9D9"),) * 4)

    # 기준 모집단(필터)
    base_mask = (
        (df_all_["일별"].dt.date >= d_start_)
        & (df_all_["일별"].dt.date <= d_end_)
        & (df_all_["country"] == "US")
        & (df_all_["objective"] == "Traffic")
    )
    base = df_all_[base_mask]
    raw_total = len(df_all_)
    pass_total = len(base)

    wb = Workbook()
    ws = wb.active
    ws.title = "우수 소재 분석"

    # 컬럼 너비
    for i, (_, w, _) in enumerate(_XLSX_COLS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 메타데이터
    ws.cell(row=1, column=1,
            value=f"품목별 상위 소재 (필터: US · Traffic · 지출 ≥ ₩100,000) — 기간 {d_start_} ~ {d_end_}").font = TITLE_FONT
    ws.cell(row=2, column=1,
            value=f"소스 '2026 RD' raw {raw_total:,}행 → US·Traffic 통과 {pass_total:,}행 → (광고 이름) 단위 집계, 환율 1 USD = ₩{fx_rate_:,.0f}").font = SUB_FONT
    ws.cell(row=3, column=1,
            value="① CTR Top 10 ← 매체 TikTok    ② AA Total Sales ($) Top 10 ← 매체 Meta").font = SUB_FONT

    cur_row = 5

    def write_header(row: int):
        for i, (name, _, _) in enumerate(_XLSX_COLS, start=1):
            c = ws.cell(row=row, column=i, value=name)
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
            c.alignment = CENTER
            c.border = BORDER

    def write_data_row(row: int, rank: int, rec: dict):
        ctr_ratio = (rec["clicks"] / rec["impressions"]) if rec["impressions"] else 0
        roas_ratio = ((rec["sales_usd"] * fx_rate_) / rec["cost"]) if rec["cost"] else 0
        values = [
            rank,
            rec["광고 이름"],
            rec["item"],
            rec["media_"],
            rec["ad_code"],
            rec["aa_link"],
            float(rec["cost"]),
            int(rec["impressions"]),
            int(rec["clicks"]),
            ctr_ratio,
            float(rec["sales_usd"]),
            int(rec["purchases"]),
            roas_ratio,
        ]
        for i, (v, (_, _, nf)) in enumerate(zip(values, _XLSX_COLS), start=1):
            c = ws.cell(row=row, column=i, value=v)
            if nf:
                c.number_format = nf
            c.border = BORDER
            if i == 1:
                c.alignment = CENTER
            if i == 6 and v:  # AA 링크
                c.hyperlink = v
                c.font = LINK_FONT

    for it in items_to_include:
        # 통과 카운트
        tt_count = int(
            (base[(base["media"] == "TikTok") & (base["item"] == it)]
             .groupby("광고 이름")["cost"].sum() >= 100_000).sum()
        )
        meta_count = int(
            (base[(base["media"] == "Meta") & (base["item"] == it)]
             .groupby("광고 이름")["cost"].sum() >= 100_000).sum()
        )

        # 품목 헤더
        cell = ws.cell(row=cur_row, column=1,
                       value=f"■ {it}  (지출 ≥ ₩100,000 통과: TikTok {tt_count}개, Meta {meta_count}개)")
        cell.font = GROUP_FONT
        cur_row += 1

        # ① CTR Top 10 — TikTok
        ws.cell(row=cur_row, column=1, value="① CTR Top 10 — TikTok").font = SECTION_FONT
        cur_row += 1
        write_header(cur_row)
        cur_row += 1
        ctr_df = _top_creatives(
            df_all_, d_start_, d_end_,
            country="US", media="TikTok", objective="Traffic",
            min_cost=100_000, sort_col="ctr", fx_rate=fx_rate_,
            item_filter=[it],
        )
        if ctr_df.empty:
            ws.cell(row=cur_row, column=1, value="(해당 소재 없음)").font = SUB_FONT
            cur_row += 1
        else:
            for i, rec in enumerate(ctr_df.to_dict("records"), start=1):
                write_data_row(cur_row, i, rec)
                cur_row += 1
        cur_row += 1  # 빈 행

        # ② AA Total Sales Top 10 — Meta
        ws.cell(row=cur_row, column=1, value="② AA Total Sales ($) Top 10 — Meta").font = SECTION_FONT
        cur_row += 1
        write_header(cur_row)
        cur_row += 1
        sales_df = _top_creatives(
            df_all_, d_start_, d_end_,
            country="US", media="Meta", objective="Traffic",
            min_cost=100_000, sort_col="sales_usd", fx_rate=fx_rate_,
            item_filter=[it],
        )
        if sales_df.empty:
            ws.cell(row=cur_row, column=1, value="(해당 소재 없음)").font = SUB_FONT
            cur_row += 1
        else:
            for i, rec in enumerate(sales_df.to_dict("records"), start=1):
                write_data_row(cur_row, i, rec)
                cur_row += 1
        cur_row += 2  # 품목 사이 2줄 간격

    ws.freeze_panes = "A4"

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


# === Tab 6: Top Creatives ===
with tab6:
    st.markdown("### 🏆 우수 소재 Top 10 (품목 기준)")
    st.caption(
        "사이드바의 기간 필터만 적용되며, 그 외 조건은 아래에 고정됩니다. "
        "각 소재는 '광고 이름' 단위로 집계 · 지출 NET 10만원 이상만 포함."
    )

    # 품목 필터 옵션: US × Traffic × (TikTok|Meta) 조합에서 발생한 품목만 노출
    _src_mask = (
        (df_all["country"] == "US")
        & (df_all["objective"] == "Traffic")
        & (df_all["media"].isin(["TikTok", "Meta"]))
    )
    item_options = sorted(df_all.loc[_src_mask, "item"].dropna().astype(str).unique().tolist())

    sel_item = st.multiselect(
        "🧴 품목 필터",
        options=item_options,
        default=item_options,
        help="선택한 품목에 해당하는 소재만 두 Top 10 표에 표시됩니다. 비우면 전체.",
        placeholder="모든 품목",
    )
    if not sel_item:
        sel_item = item_options

    # 품목별 일괄 엑셀 다운로드
    xlsx_bytes = _build_top10_xlsx(df_all, d_start, d_end, fx_rate, sel_item)
    st.download_button(
        "⬇️ 품목별 Top 10 한 번에 엑셀로 다운로드 (xlsx)",
        data=xlsx_bytes,
        file_name=f"top10_by_item_{d_start}_{d_end}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        help="시트 2개: 'CTR Top10 (TikTok)' / 'Sales Top10 (Meta)'. 선택된 품목별로 각 10개씩 포함.",
    )

    col_a, col_b = st.columns(2)

    with col_a:
        st.markdown("#### 🎯 CTR Top 10")
        st.caption("**조건**: 국가 = US · 매체 = TikTok · 캠페인 = Traffic · 지출(NET) ≥ ₩100,000")
        top_ctr = _top_creatives(
            df_all, d_start, d_end,
            country="US", media="TikTok", objective="Traffic",
            min_cost=100_000, sort_col="ctr", fx_rate=fx_rate,
            item_filter=sel_item,
        )
        _render_top(top_ctr)

    with col_b:
        st.markdown("#### 💰 AA Total Sales Top 10")
        st.caption("**조건**: 국가 = US · 매체 = Meta · 캠페인 = Traffic · 지출(NET) ≥ ₩100,000")
        top_sales = _top_creatives(
            df_all, d_start, d_end,
            country="US", media="Meta", objective="Traffic",
            min_cost=100_000, sort_col="sales_usd", fx_rate=fx_rate,
            item_filter=sel_item,
        )
        _render_top(top_sales)


st.markdown("---")
st.caption(
    f"데이터 행수: {len(df):,} / 전체 {len(df_all):,}  ·  "
    f"마지막 업데이트: {df_all['일별'].max().date()}  ·  "
    "지출은 KRW (NET, '비용' 컬럼 기준), 매출은 AA Total Sales(USD)를 환율로 환산. ROAS·CPC·CPM 모두 NET 기준"
)
